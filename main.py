import customtkinter as ctk
from tkinter import ttk, messagebox, StringVar, END, filedialog
import mysql.connector
import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import DB_CONFIG


class Hospital:
    def __init__(self, root):
        self.root = root
        self.root.title("Hospital Management System")
        self.root.geometry("1366x768+0+0")

        self.Nameoftablets      = StringVar()
        self.ref                = StringVar()
        self.Dose               = StringVar()
        self.NumberOfTablets    = StringVar()
        self.Lot                = StringVar()
        self.Issuedate          = StringVar()
        self.ExpDate            = StringVar()
        self.DailyDose          = StringVar()
        self.sideEffect         = StringVar()
        self.FurtherInformation = StringVar()
        self.StorageAdvice      = StringVar()
        self.DrivingUsingMachine= StringVar()
        self.HowToUseMedication = StringVar()
        self.PatientId          = StringVar()
        self.DateOfBirth        = StringVar()
        self.PatientAddress     = StringVar()
        self.nhsnumber          = StringVar()
        self.patientname        = StringVar()
        self.search_var         = StringVar()

        # ── Başlık ──────────────────────────────────────────────────────────
        title = ctk.CTkLabel(self.root, text="HOSPITAL MANAGEMENT SYSTEM",
                             font=("Times New Roman", 38, "bold"),
                             text_color="red", fg_color="white",
                             corner_radius=0, height=75)
        title.pack(side="top", fill="x")

        # ── Data Frame ───────────────────────────────────────────────────────
        data_frame = ctk.CTkFrame(self.root, corner_radius=8, border_width=2,
                                  width=1366, height=310)
        data_frame.pack(fill="x")
        data_frame.pack_propagate(False)

        left_frame = ctk.CTkFrame(data_frame, corner_radius=6, border_width=1,
                                  width=870, height=290)
        left_frame.place(x=6, y=8)
        left_frame.grid_propagate(False)

        ctk.CTkLabel(left_frame, text="Patient Information",
                     font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=6,
                                                       sticky="w", padx=10, pady=(4,0))

        right_frame = ctk.CTkFrame(data_frame, corner_radius=6, border_width=1,
                                   width=472, height=290)
        right_frame.place(x=884, y=8)
        right_frame.pack_propagate(False)

        ctk.CTkLabel(right_frame, text="Prescription",
                     font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=(4,0))

        # ── Sol çerçeve widget'ları ──────────────────────────────────────────
        fl = ("Arial", 11, "bold")
        fe = ("Arial", 11)
        ew = 130

        def lbl(parent, text, row, col, px=(10,2)):
            ctk.CTkLabel(parent, text=text, font=fl).grid(
                row=row, column=col, sticky="w", padx=px, pady=2)

        def ent(parent, var, row, col, px=(2,8)):
            e = ctk.CTkEntry(parent, textvariable=var, font=fe, width=ew)
            e.grid(row=row, column=col, sticky="w", padx=px, pady=2)
            return e

        # Sütun 0-1
        lbl(left_frame, "Name of Tablet:", 1, 0)
        ctk.CTkComboBox(left_frame, variable=self.Nameoftablets,
                        values=["Nice","Corona Vaccine","Acetaminophen",
                                "Adderall","Amlodipine","Ativan"],
                        font=fe, width=ew).grid(row=1, column=1, sticky="w", padx=(2,8), pady=2)

        lbl(left_frame, "Reference No:",  2, 0); ent(left_frame, self.ref,              2, 1)
        lbl(left_frame, "Dose:",          3, 0); ent(left_frame, self.Dose,             3, 1)
        lbl(left_frame, "No Of Tablets:", 4, 0); ent(left_frame, self.NumberOfTablets,  4, 1)
        lbl(left_frame, "Lot:",           5, 0); ent(left_frame, self.Lot,              5, 1)
        lbl(left_frame, "Issue Date:",    6, 0); ent(left_frame, self.Issuedate,        6, 1)
        lbl(left_frame, "Exp Date:",      7, 0); ent(left_frame, self.ExpDate,          7, 1)

        # Sütun 2-3
        lbl(left_frame, "Daily Dose:",     1, 2, px=(16,2)); ent(left_frame, self.DailyDose,          1, 3)
        lbl(left_frame, "Side Effect:",    2, 2, px=(16,2)); ent(left_frame, self.sideEffect,         2, 3)
        lbl(left_frame, "Further Info:",   3, 2, px=(16,2)); ent(left_frame, self.FurtherInformation, 3, 3)
        lbl(left_frame, "Blood Pressure:", 4, 2, px=(16,2)); ent(left_frame, self.DrivingUsingMachine,4, 3)
        lbl(left_frame, "Storage Advice:", 5, 2, px=(16,2)); ent(left_frame, self.StorageAdvice,      5, 3)
        lbl(left_frame, "Medication:",     6, 2, px=(16,2)); ent(left_frame, self.HowToUseMedication, 6, 3)
        lbl(left_frame, "Patient Id:",     7, 2, px=(16,2)); ent(left_frame, self.PatientId,          7, 3)
        lbl(left_frame, "NHS Number:",     8, 2, px=(16,2)); ent(left_frame, self.nhsnumber,          8, 3)

        # Sütun 4-5
        lbl(left_frame, "Patient Name:",    1, 4, px=(16,2)); ent(left_frame, self.patientname, 1, 5)
        lbl(left_frame, "Date Of Birth:",   2, 4, px=(16,2)); ent(left_frame, self.DateOfBirth,  2, 5)
        lbl(left_frame, "Patient Address:", 3, 4, px=(16,2))
        self.txtPatientAdress = ctk.CTkTextbox(left_frame, font=fe, width=ew, height=108)
        self.txtPatientAdress.grid(row=3, column=5, rowspan=5, sticky="w", padx=(2,8), pady=2)
        self.txtPatientAdress.bind("<KeyRelease>",
            lambda e: self.PatientAddress.set(self.txtPatientAdress.get("1.0", "end-1c")))

        # Sağ çerçeve — Prescription
        self.txtPrescription = ctk.CTkTextbox(right_frame, font=fe, width=448, height=252)
        self.txtPrescription.pack(padx=6, pady=4)

        # ── Buton Çerçevesi ─────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(self.root, corner_radius=6, border_width=1,
                                 width=1366, height=52)
        btn_frame.pack(fill="x")
        btn_frame.pack_propagate(False)

        btn_cfg = dict(font=("Arial", 11, "bold"), height=34,
                       fg_color="white", text_color="green",
                       border_width=1, border_color="green",
                       hover_color="#e6f4ea", corner_radius=6)

        ctk.CTkButton(btn_frame, text="Prescription",      command=self.Prescription,      width=170, **btn_cfg).grid(row=0, column=0, padx=5, pady=8)
        ctk.CTkButton(btn_frame, text="Prescription Data", command=self.iPrescriptionData, width=170, **btn_cfg).grid(row=0, column=1, padx=5, pady=8)
        ctk.CTkButton(btn_frame, text="Update",            command=self.update,            width=170, **btn_cfg).grid(row=0, column=2, padx=5, pady=8)
        ctk.CTkButton(btn_frame, text="Delete",            command=self.idelete,           width=170, **btn_cfg).grid(row=0, column=3, padx=5, pady=8)
        ctk.CTkButton(btn_frame, text="Clear",             command=self.clear,             width=170, **btn_cfg).grid(row=0, column=4, padx=5, pady=8)
        ctk.CTkButton(btn_frame, text="Exit",              command=self.iExit,             width=170, **btn_cfg).grid(row=0, column=5, padx=5, pady=8)

        # ── Arama & Export Çerçevesi ─────────────────────────────────────────
        search_frame = ctk.CTkFrame(self.root, corner_radius=6, border_width=1,
                                    width=1366, height=46)
        search_frame.pack(fill="x")
        search_frame.pack_propagate(False)

        ctk.CTkLabel(search_frame, text="Search by Reference No:",
                     font=("Arial", 11, "bold")).grid(row=0, column=0, padx=(12,4), pady=8)
        ctk.CTkEntry(search_frame, textvariable=self.search_var,
                     font=("Arial", 11), width=200).grid(row=0, column=1, padx=(0,4), pady=8)
        ctk.CTkButton(search_frame, text="Search",         command=self.search_data,  width=100, **btn_cfg).grid(row=0, column=2, padx=(0,4), pady=8)
        ctk.CTkButton(search_frame, text="Show All",       command=self.fetch_data,   width=100, **btn_cfg).grid(row=0, column=3, padx=(0,20), pady=8)
        ctk.CTkButton(search_frame, text="Export to CSV",  command=self.export_csv,   width=130, **btn_cfg).grid(row=0, column=4, padx=(0,4), pady=8)
        ctk.CTkButton(search_frame, text="Export to Excel",command=self.export_excel, width=130, **btn_cfg).grid(row=0, column=5, padx=(0,4), pady=8)

        # ── Tablo Çerçevesi ─────────────────────────────────────────────────
        table_frame = ctk.CTkFrame(self.root, corner_radius=6, border_width=1)
        table_frame.pack(fill="both", expand=True, pady=(0,2))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                         background="#f9f9f9", foreground="#1a1a1a",
                         rowheight=26, fieldbackground="#f9f9f9",
                         font=("Arial", 10))
        style.configure("Treeview.Heading",
                         background="#2d6a4f", foreground="white",
                         font=("Arial", 10, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#b7e4c7")])

        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
        scroll_y = ttk.Scrollbar(table_frame, orient="vertical")

        cols = ("nameoftable","ref","dose","nooftablets","lot","issuedate",
                "expdate","dailydose","storage","nhsnumber","pname","dob","address")

        self.hospital_table = ttk.Treeview(table_frame, columns=cols,
                                           xscrollcommand=scroll_x.set,
                                           yscrollcommand=scroll_y.set,
                                           show="headings")

        scroll_x.pack(side="bottom", fill="x")
        scroll_y.pack(side="right",  fill="y")
        scroll_x.config(command=self.hospital_table.xview)
        scroll_y.config(command=self.hospital_table.yview)

        headings = ["Name Of Tablet","Reference No","Dose","No Of Tablets","Lot",
                    "Issue Date","Exp Date","Daily Dose","Storage","NHS Number",
                    "Patient Name","Date Of Birth","Address"]

        for col, heading in zip(cols, headings):
            self.hospital_table.heading(col, text=heading)
            self.hospital_table.column(col, width=100, anchor="center")

        self.hospital_table.pack(fill="both", expand=True)
        self.hospital_table.bind("<ButtonRelease-1>", self.get_cursor)
        self.fetch_data()

    # ── Yardımcılar ──────────────────────────────────────────────────────────
    def _connect(self):
        return mysql.connector.connect(**DB_CONFIG)

    def validate_dates(self):
        formats = ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]

        def parse(text, field):
            for fmt in formats:
                try:
                    return datetime.datetime.strptime(text, fmt)
                except ValueError:
                    continue
            messagebox.showerror("Invalid Date",
                f"{field} must be in DD/MM/YYYY, DD-MM-YYYY or DD.MM.YYYY format.")
            return None

        issue = parse(self.Issuedate.get(),   "Issue Date")
        if issue is None: return False
        exp   = parse(self.ExpDate.get(),     "Exp Date")
        if exp   is None: return False
        dob   = parse(self.DateOfBirth.get(), "Date Of Birth")
        if dob   is None: return False

        if exp <= issue:
            messagebox.showerror("Invalid Date", "Expiry Date must be later than Issue Date.")
            return False
        if dob >= datetime.datetime.today():
            messagebox.showerror("Invalid Date", "Date Of Birth cannot be today or a future date.")
            return False
        return True

    def _get_table_rows(self):
        return [self.hospital_table.item(i)["values"]
                for i in self.hospital_table.get_children()]

    HEADERS = ["Name Of Tablet","Reference No","Dose","No Of Tablets","Lot",
               "Issue Date","Exp Date","Daily Dose","Storage","NHS Number",
               "Patient Name","Date Of Birth","Address"]

    # ── Fonksiyonlar ─────────────────────────────────────────────────────────
    def get_cursor(self, event=" "):
        cursor_row = self.hospital_table.focus()
        content    = self.hospital_table.item(cursor_row)
        row        = content["values"]
        if not row:
            return
        self.Nameoftablets.set(row[0])
        self.ref.set(row[1])
        self.Dose.set(row[2])
        self.NumberOfTablets.set(row[3])
        self.Lot.set(row[4])
        self.Issuedate.set(row[5])
        self.ExpDate.set(row[6])
        self.DailyDose.set(row[7])
        self.StorageAdvice.set(row[8])
        self.nhsnumber.set(row[9])
        self.patientname.set(row[10])
        self.DateOfBirth.set(row[11])
        self.PatientAddress.set(row[12])
        self.txtPatientAdress.delete("1.0", END)
        self.txtPatientAdress.insert("1.0", row[12])

    def Prescription(self):
        self.txtPrescription.insert(END, "Name of Tablets:\t\t"    + self.Nameoftablets.get()       + "\n")
        self.txtPrescription.insert(END, "Reference No:\t\t"        + self.ref.get()                 + "\n")
        self.txtPrescription.insert(END, "Dose:\t\t\t"              + self.Dose.get()                + "\n")
        self.txtPrescription.insert(END, "Number Of Tablets:\t\t"   + self.NumberOfTablets.get()     + "\n")
        self.txtPrescription.insert(END, "Lot:\t\t\t"               + self.Lot.get()                 + "\n")
        self.txtPrescription.insert(END, "Issue Date:\t\t"          + self.Issuedate.get()           + "\n")
        self.txtPrescription.insert(END, "Exp Date:\t\t"            + self.ExpDate.get()             + "\n")
        self.txtPrescription.insert(END, "Daily Dose:\t\t"          + self.DailyDose.get()           + "\n")
        self.txtPrescription.insert(END, "Side Effect:\t\t"         + self.sideEffect.get()          + "\n")
        self.txtPrescription.insert(END, "Further Information:\t\t" + self.FurtherInformation.get()  + "\n")
        self.txtPrescription.insert(END, "Storage Advice:\t\t"      + self.StorageAdvice.get()       + "\n")
        self.txtPrescription.insert(END, "Driving/Machine:\t\t"     + self.DrivingUsingMachine.get() + "\n")
        self.txtPrescription.insert(END, "Patient Id:\t\t"          + self.PatientId.get()           + "\n")
        self.txtPrescription.insert(END, "NHS Number:\t\t"          + self.nhsnumber.get()           + "\n")
        self.txtPrescription.insert(END, "Patient Name:\t\t"        + self.patientname.get()         + "\n")
        self.txtPrescription.insert(END, "Date Of Birth:\t\t"       + self.DateOfBirth.get()         + "\n")
        self.txtPrescription.insert(END, "Patient Address:\t\t"     + self.PatientAddress.get()      + "\n")

    def iPrescriptionData(self):
        if self.Nameoftablets.get() == "" or self.ref.get() == "":
            messagebox.showerror("Error", "All fields are required")
            return
        if not self.validate_dates():
            return
        conn = None
        try:
            conn = self._connect()
            cur  = conn.cursor()
            cur.execute("""
                INSERT INTO hospital
                (Nameoftablets, Reference_No, dose, Numbersoftablets, lot,
                 issuedate, expdate, dailydose, storage, nhsnumber,
                 patientname, DOB, patientaddress)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                self.Nameoftablets.get(), self.ref.get(),
                self.Dose.get(), self.NumberOfTablets.get(), self.Lot.get(),
                self.Issuedate.get(), self.ExpDate.get(), self.DailyDose.get(),
                self.StorageAdvice.get(), self.nhsnumber.get(),
                self.patientname.get(), self.DateOfBirth.get(), self.PatientAddress.get()
            ))
            conn.commit()
            self.fetch_data()
            messagebox.showinfo("Success", "Record has been inserted")
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", str(err))
        finally:
            if conn and conn.is_connected(): conn.close()

    def update(self):
        if self.ref.get() == "":
            messagebox.showerror("Error", "Please select a record from the table to update")
            return
        if not self.validate_dates():
            return
        conn = None
        try:
            conn = self._connect()
            cur  = conn.cursor()
            cur.execute("""
                UPDATE hospital SET
                    Nameoftablets=%s, dose=%s, Numbersoftablets=%s, lot=%s,
                    issuedate=%s, expdate=%s, dailydose=%s, storage=%s,
                    nhsnumber=%s, patientname=%s, DOB=%s, patientaddress=%s
                WHERE Reference_No=%s
            """, (
                self.Nameoftablets.get(), self.Dose.get(),
                self.NumberOfTablets.get(), self.Lot.get(),
                self.Issuedate.get(), self.ExpDate.get(),
                self.DailyDose.get(), self.StorageAdvice.get(),
                self.nhsnumber.get(), self.patientname.get(),
                self.DateOfBirth.get(), self.PatientAddress.get(),
                self.ref.get()
            ))
            conn.commit()
            self.fetch_data()
            messagebox.showinfo("Success", "Record has been updated")
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", str(err))
        finally:
            if conn and conn.is_connected(): conn.close()

    def idelete(self):
        if self.ref.get() == "":
            messagebox.showerror("Error", "Please select a record from the table to delete")
            return
        conn = None
        try:
            conn = self._connect()
            cur  = conn.cursor()
            cur.execute("DELETE FROM hospital WHERE Reference_No=%s", (self.ref.get(),))
            conn.commit()
            self.fetch_data()
            messagebox.showinfo("Delete", "Patient has been deleted successfully")
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", str(err))
        finally:
            if conn and conn.is_connected(): conn.close()

    def clear(self):
        for var in [self.Nameoftablets, self.ref, self.Dose, self.NumberOfTablets,
                    self.Lot, self.Issuedate, self.ExpDate, self.DailyDose,
                    self.sideEffect, self.FurtherInformation, self.StorageAdvice,
                    self.DrivingUsingMachine, self.HowToUseMedication, self.PatientId,
                    self.nhsnumber, self.patientname, self.DateOfBirth, self.PatientAddress]:
            var.set("")
        self.txtPatientAdress.delete("1.0", END)
        self.txtPrescription.delete("1.0", END)

    def iExit(self):
        if messagebox.askyesno("Hospital Management System", "Confirm you want to exit"):
            self.root.destroy()

    def search_data(self):
        if self.search_var.get() == "":
            messagebox.showerror("Error", "Please enter a Reference No to search")
            return
        conn = None
        try:
            conn = self._connect()
            cur  = conn.cursor()
            cur.execute("SELECT * FROM hospital WHERE Reference_No LIKE %s",
                        ("%" + self.search_var.get() + "%",))
            rows = cur.fetchall()
            self.hospital_table.delete(*self.hospital_table.get_children())
            for i in rows:
                self.hospital_table.insert("", END, values=i)
            if not rows:
                messagebox.showinfo("Search Result", "No records found for the given Reference No")
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", str(err))
        finally:
            if conn and conn.is_connected(): conn.close()

    def export_csv(self):
        rows = self._get_table_rows()
        if not rows:
            messagebox.showwarning("Export", "There is no data to export")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="hospital_records.csv",
            title="Save CSV File"
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self.HEADERS)
                writer.writerows(rows)
            messagebox.showinfo("Export Successful", f"Data exported to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def export_excel(self):
        rows = self._get_table_rows()
        if not rows:
            messagebox.showwarning("Export", "There is no data to export")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="hospital_records.xlsx",
            title="Save Excel File"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hospital Records"

            header_font  = Font(bold=True, color="FFFFFF")
            header_fill  = PatternFill("solid", fgColor="2D6A4F")
            header_align = Alignment(horizontal="center", vertical="center")

            for col_num, header in enumerate(self.HEADERS, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align

            for row_num, row in enumerate(rows, start=2):
                for col_num, value in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col_num, value=value)

            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            messagebox.showinfo("Export Successful", f"Data exported to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def fetch_data(self):
        conn = None
        try:
            conn = self._connect()
            cur  = conn.cursor()
            cur.execute("SELECT * FROM hospital")
            rows = cur.fetchall()
            self.hospital_table.delete(*self.hospital_table.get_children())
            for i in rows:
                self.hospital_table.insert("", END, values=i)
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", str(err))
        finally:
            if conn and conn.is_connected(): conn.close()


root = ctk.CTk()
ob = Hospital(root)
root.mainloop()